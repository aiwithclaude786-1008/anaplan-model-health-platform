# app/reports/excel_report.py -- master spec section 20.
# ============================================================
# 10-sheet consultant workbook, generalized from the real
# Tridant Module Sizing / Big Line Items / Findings report
# structure -- never hardcoded to a specific client. Every sheet
# is produced from the same AnalysisResult every UI page reads.
# ============================================================
from __future__ import annotations

import io

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.analysis.pipeline import AnalysisResult
from app.analysis.size_analysis import cells_to_gb
from app.rules.registry import get_all_rules

_HEADER_FILL = PatternFill(start_color="00ADEF", end_color="00ADEF", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def _style_header(ws):
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    for i, col in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(i)].width = min(60, max(10, max_len + 2))


def build_excel_report(result: AnalysisResult, client_name: str = "", model_label: str = "") -> bytes:
    size = result.size
    health = result.health

    exec_rows = [
        {"Metric": "Client", "Value": client_name or "N/A"},
        {"Metric": "Model", "Value": model_label or "N/A"},
        {"Metric": "Overall Health Score", "Value": health.overall},
        {"Metric": "Health Band", "Value": health.band},
        {"Metric": "Total Cells", "Value": size.total_cells if size.cell_count_available else "N/A"},
        {"Metric": "Estimated Model Size (GB)", "Value": round(cells_to_gb(size.total_cells), 2) if size.cell_count_available else "N/A"},
        {"Metric": "Modules", "Value": size.modules_count},
        {"Metric": "Line Items", "Value": size.line_items_count},
        {"Metric": f"Top {size.top_n_actual} Modules' Share", "Value": f"{size.top_n_pct:.1f}%" if size.top_n_pct is not None else "N/A"},
        {"Metric": "Data Quality Score", "Value": result.data_quality.score},
        {"Metric": "Critical Findings", "Value": sum(1 for f in result.findings if f.severity == "critical")},
        {"Metric": "High Findings", "Value": sum(1 for f in result.findings if f.severity == "high")},
    ]
    for d in health.dimensions:
        exec_rows.append({"Metric": d.label, "Value": d.score})
    df_exec = pd.DataFrame(exec_rows)

    df_size = pd.DataFrame([{
        "Rank": r.rank, "Module": r.module, "Cell Count": r.cell_count, "% of Model": round(r.pct_of_model * 100, 2),
        "Cumulative %": round(r.cumulative_pct * 100, 2), "Line Items": r.line_items,
        "Avg Cells/Item": round(r.avg_cells_per_item, 1), "Module Type": r.module_type,
        "Status": r.status, "Primary Lever": r.primary_lever,
    } for r in size.module_rows]) if size.cell_count_available else pd.DataFrame()

    feats = result.feats
    df_module_analysis = feats.groupby("module").agg(
        line_items=("line_item", "count"),
        avg_formula_length=("formula_length", "mean"),
        avg_func_density=("func_density_count", "mean"),
        avg_impact_score=("impact_score", "mean"),
    ).reset_index().rename(columns={"module": "Module"}).sort_values("avg_impact_score", ascending=False)

    df_formula = feats[["module", "line_item", "formula", "formula_length", "count_if", "count_lookup",
                         "count_sum", "count_select", "func_density_count", "impact_score"]].rename(columns={
        "module": "Module", "line_item": "Line Item", "formula": "Formula", "formula_length": "Length",
        "count_if": "IF", "count_lookup": "LOOKUP", "count_sum": "SUM", "count_select": "SELECT",
        "func_density_count": "Function Density", "impact_score": "Formula Impact Score",
    }).sort_values("Formula Impact Score", ascending=False)

    df_performance = pd.DataFrame([{
        "Module": p.module, "Line Item": p.line_item, "Complexity": p.complexity,
        "Cell Count": p.cell_count, "Quadrant": p.quadrant, "Impact Score": p.impact_score,
    } for p in result.hotspots]) if result.hotspots else pd.DataFrame()

    dim = result.dimensionality
    df_dimensionality = pd.DataFrame([{
        "Module": r.module, "Distinct Applies To": r.distinct_applies_to,
        "Subsidiary-view items": r.subsidiary_view_items, "Full-calendar items": r.full_calendar_items,
        "Full-calendar cells": r.full_calendar_cells,
    } for r in dim.module_rows]) if dim.waste_score is not None else pd.DataFrame([{"Note": dim.note}])

    df_findings = pd.DataFrame([f.to_dict() for f in result.findings])

    df_recommendations = pd.DataFrame([{
        "Priority": o.priority, "Module": o.module, "Issue": o.issue, "Recommended Action": o.recommended_action,
        "Expected Benefit": o.expected_benefit, "Confidence": o.confidence, "Effort": o.effort,
    } for o in result.top_opportunities])

    df_backlog = pd.DataFrame([{
        "Priority": o.priority, "Module": o.module, "Item(s)": o.line_item, "Issue": o.issue,
        "Current Impact": o.current_impact, "Recommended Action": o.recommended_action,
        "Expected Benefit": o.expected_benefit, "Confidence": o.confidence, "Effort": o.effort,
        "Validation Required": o.validation_required, "Severity": o.severity,
    } for o in (result.top_opportunities + result.size_opportunities)])

    rules = get_all_rules()
    df_rules = pd.DataFrame([{
        "Rule ID": r.rule_id, "Name": r.name, "Category": r.category, "Severity": r.severity,
        "Description": r.description, "Recommendation": r.recommendation, "Confidence": r.confidence,
        "Affects Size": r.affects_size, "Affects Performance": r.affects_performance,
    } for r in rules])

    sheets = {
        "1. Exec Summary": df_exec,
        "2. Module Sizing": df_size,
        "3. Module Analysis": df_module_analysis,
        "4. Formula Analysis": df_formula.head(2000),
        "5. Performance Hotspots": df_performance,
        "6. Dimensionality": df_dimensionality,
        "7. Findings": df_findings.head(5000),
        "8. Recommendations": df_recommendations,
        "9. Optimization Backlog": df_backlog,
        "10. Rules Reference": df_rules,
    }

    buffer = io.BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        for name, df in sheets.items():
            (df if not df.empty else pd.DataFrame([{"Note": "No data for this section."}])).to_excel(
                writer, sheet_name=name[:31], index=False
            )
        for ws in writer.book.worksheets:
            _style_header(ws)
    return buffer.getvalue()
